from flask import Flask, jsonify, request, render_template_string, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import json
import os
import uuid
import io

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Google Sheets Integration
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# Google Sheets setup
def get_google_client():
    """Get gspread client from environment variable"""
    if not GSPREAD_AVAILABLE:
        return None

    import os
    import json

    creds_json = os.environ.get('GOOGLE_CREDENTIALS')
    if not creds_json:
        print("WARNING: GOOGLE_CREDENTIALS not set")
        return None

    try:
        # Parse the JSON credentials from env variable
        creds_dict = json.loads(creds_json)

        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]

        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(credentials)
    except Exception as e:
        print(f"ERROR connecting to Google Sheets: {e}")
        return None

def get_or_create_sheet(sheet_name, worksheet_name):
    """Get or create a worksheet in the spreadsheet"""
    gc = get_google_client()
    if not gc:
        return None

    try:
        # Try to open existing spreadsheet
        spreadsheet = gc.open(sheet_name)
    except gspread.exceptions.SpreadsheetNotFound:
        # Create new spreadsheet
        spreadsheet = gc.create(sheet_name)
        # Share with yourself (optional)
        # spreadsheet.share('your-email@gmail.com', perm_type='user', role='writer')

    try:
        worksheet = spreadsheet.worksheet(worksheet_name)
    except gspread.exceptions.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)

    return worksheet

def sync_inventory_to_sheets():
    """Sync current inventory to Google Sheets"""
    try:
        ws = get_or_create_sheet("Gerrit POS Data", "Inventory")
        if not ws:
            print("ERROR: Could not get Inventory worksheet")
            return False

        # Clear existing data
        ws.clear()

        # Headers
        headers = ['Product', 'Stock', 'Price', 'Value', 'Last Updated']
        ws.append_row(headers)

        # Data
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for name, data in inventory_data.items():
            value = data['stock'] * data['price']
            ws.append_row([name, data['stock'], data['price'], value, now])

        print(f"SUCCESS: Synced {len(inventory_data)} items to Inventory sheet")
        return True
    except Exception as e:
        print(f"ERROR syncing inventory: {e}")
        return False

def sync_sales_to_sheets():
    """Sync sales to Google Sheets"""
    try:
        ws = get_or_create_sheet("Gerrit POS Data", "Sales")
        if not ws:
            print("ERROR: Could not get Sales worksheet")
            return False

        ws.clear()
        headers = ['Date', 'Transaction ID', 'Product', 'Quantity', 'Unit Price', 'Total', 'M-Pesa', 'Cash', 'Debt']
        ws.append_row(headers)

        for sale in sales_data:
            ws.append_row([
                sale['date'],
                sale.get('transactionId', '-'),
                sale['product'],
                sale['quantity'],
                sale['unitPrice'],
                sale['total'],
                sale.get('mpesa', 0),
                sale.get('cash', 0),
                sale.get('debt', 0)
            ])

        print(f"SUCCESS: Synced {len(sales_data)} sales to Sales sheet")
        return True
    except Exception as e:
        print(f"ERROR syncing sales: {e}")
        return False

def sync_restocks_to_sheets():
    """Sync restocks to Google Sheets"""
    try:
        ws = get_or_create_sheet("Gerrit POS Data", "Restocks")
        if not ws:
            print("ERROR: Could not get Restocks worksheet")
            return False

        ws.clear()
        headers = ['Date', 'Product', 'Quantity', 'Unit Price', 'Total']
        ws.append_row(headers)

        for restock in restocks_data:
            ws.append_row([
                restock['date'],
                restock['product'],
                restock['quantity'],
                restock['unitPrice'],
                restock['total']
            ])

        print(f"SUCCESS: Synced {len(restocks_data)} restocks to Restocks sheet")
        return True
    except Exception as e:
        print(f"ERROR syncing restocks: {e}")
        return False

def load_inventory_from_sheets():
    """Load inventory from Google Sheets (if available)"""
    global inventory_data

    ws = get_or_create_sheet("Gerrit POS Data", "Inventory")
    if not ws:
        return False

    try:
        records = ws.get_all_records()
        if records:
            new_inventory = {}
            for row in records:
                product = row.get('Product', '')
                if product:
                    new_inventory[product] = {
                        'stock': float(row.get('Stock', 0)),
                        'price': float(row.get('Price', 0))
                    }
            if new_inventory:
                inventory_data = new_inventory
                save_inventory()
                return True
    except Exception as e:
        print(f"Error loading from sheets: {e}")

    return False

app = Flask(__name__)
CORS(app)

INVENTORY_FILE = '/tmp/inventory.json'
SALES_FILE = '/tmp/sales.json'
RESTOCKS_FILE = '/tmp/restocks.json'

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'padmin123'

active_tokens = set()

initial_inventory = {"all seasons 250ml": {"stock": 4.0, "price": 450.0}, "all seasons 375ml": {"stock": 0.0, "price": 650.0}, "all seasons 750ml": {"stock": 2.0, "price": 1300.0}, "Alter Wine": {"stock": 1.0, "price": 1400.0}, "Asconi": {"stock": 1.0, "price": 1950.0}, "Best Gin  250ml": {"stock": 0.0, "price": 320.0}, "Best Gin  750ml": {"stock": 0.0, "price": 900.0}, "Best vodka 250ml": {"stock": 0.0, "price": 300.0}, "Best vodka 750ml": {"stock": 3.0, "price": 850.0}, "Best whisky 750ml": {"stock": 0.0, "price": 1150.0}, "Best whisky 250ml": {"stock": 0.0, "price": 300.0}, "Blue ice": {"stock": 4.0, "price": 180.0}, "Bacardi": {"stock": 1.0, "price": 2700.0}, "balozi": {"stock": 6.0, "price": 240.0}, "Bombay Saphire": {"stock": 1.0, "price": 3450.0}, "Beefeater London 750ml": {"stock": 1.0, "price": 2350.0}, "beefeater pink 750ml": {"stock": 3.0, "price": 3150.0}, "beefeater pink 1litre": {"stock": 1.0, "price": 3150.0}, "beefeater london gin 1000ml": {"stock": 1.0, "price": 2850.0}, "black $ white 1000ml": {"stock": 0.0, "price": 1850.0}, "black $ white 750ml": {"stock": 0.0, "price": 1400.0}, "black $ white 350ml": {"stock": 1.0, "price": 730.0}, "bond 7 750ml": {"stock": 2.0, "price": 1600.0}, "bond 7 250ml": {"stock": 3.0, "price": 500.0}, "Black Label 1000ml": {"stock": 0.0, "price": 4500.0}, "Black Label 750ml": {"stock": 1.0, "price": 3700.0}, "black label 350ml": {"stock": 2.0, "price": 2100.0}, "Back label 250ml": {"stock": 2.0, "price": 0.0}, "black $ white 250ml": {"stock": 0.0, "price": 500.0}, "black bird": {"stock": 1.0, "price": 1100.0}, "captain morgan 1 litre (spiced)": {"stock": 1.0, "price": 2800.0}, "captain morgan 750ml": {"stock": 4.0, "price": 1150.0}, "captain morgan 250ml": {"stock": 4.0, "price": 400.0}, "chrome gin 750ml": {"stock": 3.0, "price": 700.0}, "chrome gin 250ml": {"stock": 6.0, "price": 250.0}, "chrome vodka 750ml": {"stock": 3.0, "price": 700.0}, "chrome vodka 250ml": {"stock": 7.0, "price": 250.0}, "circo": {"stock": 1.0, "price": 4500.0}, "chivas regal": {"stock": 1.0, "price": 4400.0}, "caprice sweet red": {"stock": 0.0, "price": 950.0}, "caprice white": {"stock": 1.0, "price": 950.0}, "crazy cock 750ml": {"stock": 0.0, "price": 1150.0}, "crazy cock 350ml": {"stock": 0.0, "price": 650.0}, "crazy cock 250ml": {"stock": 1.0, "price": 450.0}, "caribia gin 750ml": {"stock": 2.0, "price": 870.0}, "caribia gin 350ml": {"stock": 0.0, "price": 0.0}, "caribia gin 250ml": {"stock": 4.0, "price": 300.0}, "county 750ml": {"stock": 1.0, "price": 800.0}, "county 250ml": {"stock": 4.0, "price": 300.0}, "Camino tequlla ": {"stock": 0.0, "price": 2600.0}, "clubman 750ml ": {"stock": 1.0, "price": 900.0}, "cellar cask 5litre": {"stock": 1.0, "price": 4500.0}, "cellar cask Red 750ml": {"stock": 1.0, "price": 1050.0}, "chamdor Red": {"stock": 1.0, "price": 900.0}, "desperado": {"stock": 2.0, "price": 350.0}, "drostdy sweet red": {"stock": 4.0, "price": 1150.0}, "drostdy Sweet white": {"stock": 1.0, "price": 1150.0}, "delush Red": {"stock": 1.0, "price": 1000.0}, "eristoff": {"stock": 1.0, "price": 1400.0}, "first choice 750ml": {"stock": 3.0, "price": 800.0}, "famous grouse 1litre": {"stock": 1.0, "price": 2800.0}, "famous grouse 750ml": {"stock": 1.0, "price": 2150.0}, "faxe": {"stock": 5.0, "price": 340.0}, "gordons pink 1litre": {"stock": 2.0, "price": 2800.0}, "gordons pink 750ml": {"stock": 1.0, "price": 2300.0}, "gordons original 1litre": {"stock": 0.0, "price": 2900.0}, "gordons original 750ml": {"stock": 0.0, "price": 2400.0}, "gordons lemon 750ml": {"stock": 1.0, "price": 2300.0}, "gordons orange": {"stock": 0.0, "price": 2300.0}, "glenfiddich": {"stock": 2.0, "price": 7400.0}, "gilbeys pink 750ml": {"stock": 1.0, "price": 1600.0}, "gilbeys pink 350ml": {"stock": 1.0, "price": 710.0}, "gilbeys pink 250ml": {"stock": 1.0, "price": 500.0}, "gilbeys original 750ml": {"stock": 1.0, "price": 1600.0}, "gilgeys original 350ml": {"stock": 1.0, "price": 710.0}, "gilbeys 250ml": {"stock": 2.0, "price": 500.0}, "gibsons 750ml": {"stock": 0.0, "price": 1600.0}, "gibsons 350ml": {"stock": 0.0, "price": 0.0}, "gibsons 250ml": {"stock": 0.0, "price": 0.0}, "glen silver": {"stock": 2.0, "price": 1700.0}, "guinness": {"stock": 6.0, "price": 270.0}, "general meakins": {"stock": 3.0, "price": 270.0}, "grande france White Semi- Sweet": {"stock": 1.0, "price": 1400.0}, "hunters choice 750ml": {"stock": 4.0, "price": 1100.0}, "hunters choice 350": {"stock": 5.0, "price": 550.0}, "hunters choice 250ml": {"stock": 6.0, "price": 450.0}, "hunters gold": {"stock": 2.0, "price": 250.0}, "heinekein": {"stock": 6.0, "price": 290.0}, "henessy SP 750ML": {"stock": 1.0, "price": 6000.0}, "hendricks 1litre": {"stock": 1.0, "price": 5000.0}, "hendricks 750ml": {"stock": 3.0, "price": 4000.0}, "imperial blue 750ml": {"stock": 3.0, "price": 1050.0}, "jack daniel original": {"stock": 0.0, "price": 0.0}, "jack daniel  750ml (honey)": {"stock": 1.0, "price": 4500.0}, "jagermeister 1lite ": {"stock": 1.5, "price": 3400.0}, "jagermeister 750ml": {"stock": 1.0, "price": 2800.0}, "jameson 1lite": {"stock": 0.0, "price": 3600.0}, "jameson 750ml": {"stock": 0.0, "price": 2950.0}, "jameson 350ml": {"stock": 2.0, "price": 1450.0}, "jameson 250ml": {"stock": 0.0, "price": 0.0}, "j $ b 750ml": {"stock": 1.0, "price": 1900.0}, "jose quavo 750ml": {"stock": 0.0, "price": 2900.0}, "kc smooth 750 ml": {"stock": 2.0, "price": 900.0}, "kc smooth 350ml": {"stock": 1.0, "price": 450.0}, "kc smooth 250ml": {"stock": 3.0, "price": 320.0}, "kc pineapple 750ml": {"stock": 3.0, "price": 900.0}, "kc pineapple 350ml": {"stock": 0.0, "price": 450.0}, "kc pineapple 250ml": {"stock": 5.0, "price": 320.0}, "kc ginger $ lemon 750ml": {"stock": 3.0, "price": 900.0}, "kc ginger $ lemon 350ml": {"stock": 0.0, "price": 450.0}, "kc ginnger $ lemon 250ml": {"stock": 6.0, "price": 320.0}, "kibao voka 750ml": {"stock": 2.0, "price": 780.0}, "kibao voka 350ml": {"stock": 0.0, "price": 450.0}, "kibao vodka 250ml": {"stock": 4.0, "price": 300.0}, "kibao gin 750ml": {"stock": 2.0, "price": 750.0}, "kibao gin 350ml": {"stock": 0.0, "price": 450.0}, "kibao  gin 250ml": {"stock": 4.0, "price": 260.0}, "konyagi 750ml": {"stock": 2.0, "price": 800.0}, "konyagi 500ml": {"stock": 0.0, "price": 550.0}, "konyagi 250ml": {"stock": 5.0, "price": 300.0}, "kane 750ml": {"stock": 4.0, "price": 750.0}, "kane 250ml": {"stock": 0.0, "price": 250.0}, "k.o tonic ": {"stock": 1.0, "price": 150.0}, "k.o  bottle": {"stock": 3.0, "price": 300.0}, "leadimg warigi 750ml": {"stock": 1.0, "price": 770.0}, "mara (wine) Red": {"stock": 1.0, "price": 1300.0}, "mikado(cherry)": {"stock": 1.0, "price": 1600.0}, "mikado (pineapple)": {"stock": 1.0, "price": 1600.0}, "malibu": {"stock": 1.0, "price": 2350.0}, "ministers reserve 750ml": {"stock": 1.0, "price": 1600.0}, "monkey shoulder": {"stock": 1.0, "price": 4500.0}, "martini": {"stock": 1.0, "price": 2700.0}, "monster": {"stock": 4.0, "price": 250.0}, "manyatta (can)": {"stock": 1.0, "price": 300.0}, "manyatta ( bottle)": {"stock": 3.0, "price": 300.0}, "namaqua sweet red (wine)": {"stock": 2.0, "price": 1000.0}, "o pm vodka 750ml": {"stock": 2.0, "price": 1250.0}, "o pm vodka 350ml": {"stock": 1.0, "price": 680.0}, "o pm vodka 250ml": {"stock": 2.0, "price": 450.0}, "old monk": {"stock": 1.0, "price": 1050.0}, "oj 16%": {"stock": 6.0, "price": 400.0}, "oj 12%": {"stock": 0.0, "price": 320.0}, "Old smuggler": {"stock": 1.0, "price": 1400.0}, "paddy irish": {"stock": 0.0, "price": 1500.0}, "passport scotch": {"stock": 1.0, "price": 1350.0}, "pervack 1litre": {"stock": 1.0, "price": 1500.0}, "pervack  750ml": {"stock": 0.0, "price": 1300.0}, "penasol white wine": {"stock": 2.0, "price": 950.0}, "penasol red wine": {"stock": 1.0, "price": 950.0}, "robertson 1.5litre": {"stock": 1.0, "price": 2100.0}, "robertson 750ml": {"stock": 1.0, "price": 1200.0}, "red label 1litre ": {"stock": 0.0, "price": 2700.0}, "red label 750ml": {"stock": 0.0, "price": 2300.0}, "red label 350ml": {"stock": 0.0, "price": 1050.0}, "red label 250ml": {"stock": 0.0, "price": 700.0}, "redbull": {"stock": 0.0, "price": 230.0}, "rosso nobile (wine)": {"stock": 2.0, "price": 1500.0}, "smirnoff vodka 1litre": {"stock": 1.0, "price": 2000.0}, "smirnoff vodka 750ml": {"stock": 1.0, "price": 1600.0}, "smirnoff vodka 350ml": {"stock": 0.0, "price": 750.0}, "smirnoff vodka 250ml": {"stock": 2.0, "price": 510.0}, "smirnoff pineapple punch ": {"stock": 5.0, "price": 220.0}, "smirnoff guaranna": {"stock": 5.0, "price": 220.0}, "smirnoff black ice": {"stock": 18.0, "price": 220.0}, "sweet berry ": {"stock": 2.0, "price": 150.0}, "strumbras": {"stock": 2.0, "price": 700.0}, "savanna": {"stock": 3.0, "price": 300.0}, "southern comfort 1litre": {"stock": 1.0, "price": 2700.0}, "southern comfort 750ml": {"stock": 3.0, "price": 2400.0}, "southern comfort 350ml": {"stock": 3.0, "price": 750.0}, "sky infusion ": {"stock": 2.0, "price": 1500.0}, "star walker ": {"stock": 1.0, "price": 1500.0}, "sun chaser(wine)": {"stock": 1.0, "price": 950.0}, "singleton ": {"stock": 1.0, "price": 5700.0}, "tanqueray  1litre ( no 10)": {"stock": 1.0, "price": 6050.0}, " tanqueray 750ml ( no 10)                 ": {"stock": 1.0, "price": 5050.0}, "tanqueray  gin 1litre": {"stock": 0.0, "price": 3750.0}, "tanqueray  gin 750ml": {"stock": 1.0, "price": 2850.0}, "top secret 750ml": {"stock": 0.0, "price": 870.0}, "top secret 250ml": {"stock": 3.0, "price": 310.0}, "three barrels 750ml": {"stock": 1.0, "price": 2850.0}, "tusker lager (can)": {"stock": 4.0, "price": 240.0}, "tusker cider (can)": {"stock": 10.0, "price": 280.0}, "tusker malt (green)": {"stock": 1.0, "price": 300.0}, "versus white (wine)": {"stock": 1.0, "price": 1200.0}, "VAT 69  1 LITRE": {"stock": 0.0, "price": 2200.0}, "VAT 69  750ML": {"stock": 0.0, "price": 900.0}, "VAT 69 350ML": {"stock": 0.0, "price": 950.0}, "VAT 69 250ML": {"stock": 1.0, "price": 650.0}, "Viceroy 750ml": {"stock": 2.0, "price": 1500.0}, "viceroy 350ml": {"stock": 2.0, "price": 760.0}, "viceroy 250ml": {"stock": 2.0, "price": 520.0}, "V$A imperial ": {"stock": 1.0, "price": 900.0}, "wild turkey (bourbon)": {"stock": 1.0, "price": 4200.0}, "white cap (can)": {"stock": 6.0, "price": 270.0}, "william lawsons 1litre": {"stock": 1.0, "price": 3000.0}, "william lawsons 750ml": {"stock": 0.0, "price": 2000.0}, "william lawsons 350ml": {"stock": 2.0, "price": 1050.0}, "william lawsons 250ml": {"stock": 0.0, "price": 0.0}, "zappa black": {"stock": 1.0, "price": 1750.0}, "zappa original": {"stock": 1.0, "price": 1750.0}, "zappa blue": {"stock": 1.0, "price": 1750.0}, "# 7": {"stock": 1.0, "price": 1200.0}, "58 gin": {"stock": 1.0, "price": 1450.0}, " miniute maid ": {"stock": 8.0, "price": 160.0}, "water 500ml ": {"stock": 12.0, "price": 30.0}, "water  dasani": {"stock": 5.0, "price": 70.0}, "soda 2litre": {"stock": 6.0, "price": 200.0}, "sodalitre 1litre": {"stock": 8.0, "price": 120.0}, "dunhill double switch": {"stock": 0.0, "price": 600.0, "pieces_per_box": 20}, "dunhill single switch": {"stock": 0.0, "price": 600.0, "pieces_per_box": 20}, "pall mall (king safari)": {"stock": 0.0, "price": 300.0, "pieces_per_box": 20}, "pall mall ( menthol)": {"stock": 0.0, "price": 300.0, "pieces_per_box": 20}, "oris milano": {"stock": 0.0, "price": 400.0, "pieces_per_box": 20}, "oris menthol": {"stock": 0.0, "price": 400.0, "pieces_per_box": 20}, "rothmans red ": {"stock": 0.0, "price": 500.0, "pieces_per_box": 20}, "rothmans blue": {"stock": 0.0, "price": 500.0, "pieces_per_box": 20}, "dunhill embassy": {"stock": 0.0, "price": 600.0, "pieces_per_box": 20}, "lemonade": {"stock": 3.0, "price": 50.0}, "predator": {"stock": 3.0, "price": 70.0}}

inventory_data = {}
sales_data = []
restocks_data = []

def load_data():
    global inventory_data, sales_data, restocks_data
    if os.path.exists(INVENTORY_FILE):
        with open(INVENTORY_FILE, 'r') as f:
            inventory_data = json.load(f)
    else:
        inventory_data = json.loads(json.dumps(initial_inventory))
        save_inventory()
    if os.path.exists(SALES_FILE):
        with open(SALES_FILE, 'r') as f:
            sales_data = json.load(f)
    else:
        sales_data = []
    if os.path.exists(RESTOCKS_FILE):
        with open(RESTOCKS_FILE, 'r') as f:
            restocks_data = json.load(f)
    else:
        restocks_data = []

def save_inventory():
    with open(INVENTORY_FILE, 'w') as f:
        json.dump(inventory_data, f)

def save_sales():
    with open(SALES_FILE, 'w') as f:
        json.dump(sales_data, f)

def save_restocks():
    with open(RESTOCKS_FILE, 'w') as f:
        json.dump(restocks_data, f)

def token_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token missing'}), 401
        if token.startswith('Bearer '):
            token = token[7:]
        if token not in active_tokens:
            return jsonify({'error': 'Invalid token'}), 401
        return f(*args, **kwargs)
    return decorated

load_data()

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'app': 'running'})

@app.route('/api/sync', methods=['POST'])
@token_required
def sync_to_google_sheets():
    """Manually sync all data to Google Sheets"""
    data = request.get_json() or {}
    sync_type = data.get('type', 'all')

    results = {}

    if sync_type in ['all', 'inventory']:
        results['inventory'] = sync_inventory_to_sheets()

    if sync_type in ['all', 'sales']:
        results['sales'] = sync_sales_to_sheets()

    if sync_type in ['all', 'restocks']:
        results['restocks'] = sync_restocks_to_sheets()

    success = any(results.values())

    return jsonify({
        'success': success,
        'results': results,
        'message': 'Synced to Google Sheets' if success else 'Sync failed - check GOOGLE_CREDENTIALS'
    })

@app.route('/api/sync/status')
@token_required
def sync_status():
    """Check if Google Sheets is connected"""
    gc = get_google_client()
    creds_set = os.environ.get('GOOGLE_CREDENTIALS') is not None

    # Check if credentials are valid JSON
    creds_valid = False
    if creds_set:
        try:
            json.loads(os.environ.get('GOOGLE_CREDENTIALS'))
            creds_valid = True
        except:
            pass

    return jsonify({
        'connected': gc is not None,
        'gspread_available': GSPREAD_AVAILABLE,
        'credentials_set': creds_set,
        'credentials_valid_json': creds_valid,
        'tmp_writable': os.access('/tmp', os.W_OK),
        'inventory_file_exists': os.path.exists(INVENTORY_FILE),
        'sales_file_exists': os.path.exists(SALES_FILE),
        'tokens_file_exists': os.path.exists(TOKENS_FILE),
        'inventory_count': len(inventory_data),
        'sales_count': len(sales_data),
        'restocks_count': len(restocks_data)
    })

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        token = str(uuid.uuid4())
        active_tokens.add(token)
        return jsonify({'success': True, 'token': token, 'message': 'Login successful'})
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    token = request.headers.get('Authorization')
    if token and token.startswith('Bearer '):
        token = token[7:]
        remove_token(token)
    return jsonify({'success': True})

@app.route('/api/check-auth')
def check_auth():
    token = request.headers.get('Authorization')
    if token and token.startswith('Bearer '):
        token = token[7:]
        tokens = get_active_tokens()
        if token in tokens:
            return jsonify({'authenticated': True})
    return jsonify({'authenticated': False})

@app.route('/api/products')
def get_products():
    products = []
    for name, data in inventory_data.items():
        products.append({'name': name, 'price': data['price'], 'stock': data['stock']})
    return jsonify(products)

@app.route('/api/inventory')
@token_required
def get_inventory():
    # Add calculated fields for cigarettes
    result = {}
    for name, data in inventory_data.items():
        result[name] = dict(data)
        if 'pieces_per_box' in data:
            result[name]['piece_price'] = round(data['price'] / data['pieces_per_box'], 2)
            result[name]['total_pieces'] = int(data['stock'] * data['pieces_per_box'])
    return jsonify(result)

@app.route('/api/sale', methods=['POST'])
@token_required
def record_sale():
    data = request.get_json()
    items = data.get('items', [])
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    payments = data.get('payments', {})

    mpesa = float(payments.get('mpesa', 0) or 0)
    cash = float(payments.get('cash', 0) or 0)
    debt = float(payments.get('debt', 0) or 0)

    total_amount = sum(item['price'] * item['quantity'] for item in items)
    total_paid = mpesa + cash + debt

    if abs(total_paid - total_amount) > 0.01:
        return jsonify({'error': f'Payment total (KES {total_paid:.2f}) does not match sale total (KES {total_amount:.2f})'}), 400

    # Generate transaction ID for this sale
    transaction_id = str(uuid.uuid4())[:8]

    for item in items:
        product = item['name']
        qty = item['quantity']
        if product in inventory_data and inventory_data[product]['stock'] >= qty:
            inventory_data[product]['stock'] -= qty
            sales_data.append({
                'date': date,
                'transactionId': transaction_id,
                'product': product,
                'quantity': qty,
                'unitPrice': item['price'],
                'total': item['price'] * qty,
                'type': 'Sale',
                'mpesa': mpesa,
                'cash': cash,
                'debt': debt
            })
        else:
            return jsonify({'error': f'Insufficient stock for {product}'}), 400

    save_inventory()
    save_sales()

    # Auto-sync to Google Sheets (non-blocking)
    try:
        sync_inventory_to_sheets()
        sync_sales_to_sheets()
    except Exception as e:
        print(f"Auto-sync error: {e}")

    return jsonify({'success': True})

@app.route('/api/restock', methods=['POST'])
@token_required
def record_restock():
    data = request.get_json()
    product = data.get('product')
    qty = data.get('quantity')
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    unit_type = data.get('unit_type', 'box')  # 'box' or 'piece'

    if product not in inventory_data:
        return jsonify({'error': 'Product not found'}), 404

    # Handle cigarette piece restocking
    if unit_type == 'piece' and 'pieces_per_box' in inventory_data[product]:
        pieces_per_box = inventory_data[product]['pieces_per_box']
        boxes_to_add = qty / pieces_per_box
        inventory_data[product]['stock'] += boxes_to_add
        restocks_data.append({
            'date': date,
            'product': product,
            'quantity': qty,
            'unit_type': 'piece',
            'unitPrice': inventory_data[product]['price'] / pieces_per_box,
            'total': (inventory_data[product]['price'] / pieces_per_box) * qty,
            'type': 'Restock'
        })
    else:
        # Regular box restocking
        inventory_data[product]['stock'] += qty
        restocks_data.append({
            'date': date,
            'product': product,
            'quantity': qty,
            'unit_type': 'box',
            'unitPrice': inventory_data[product]['price'],
            'total': inventory_data[product]['price'] * qty,
            'type': 'Restock'
        })

    save_inventory()
    save_restocks()

    # Auto-sync to Google Sheets (non-blocking)
    try:
        sync_inventory_to_sheets()
        sync_restocks_to_sheets()
    except Exception as e:
        print(f"Auto-sync error: {e}")

    return jsonify({'success': True})

@app.route('/api/transactions')
@token_required
def get_transactions():
    all_transactions = sales_data + restocks_data
    return jsonify(sorted(all_transactions, key=lambda x: x['date'], reverse=True))

@app.route('/api/stats')
@token_required
def get_stats():
    today = datetime.now().strftime('%Y-%m-%d')
    today_sales = [s for s in sales_data if s['date'] == today]
    total_mpesa = sum(s.get('mpesa', 0) for s in sales_data)
    total_cash = sum(s.get('cash', 0) for s in sales_data)
    total_debt = sum(s.get('debt', 0) for s in sales_data)
    today_mpesa = sum(s.get('mpesa', 0) for s in today_sales)
    today_cash = sum(s.get('cash', 0) for s in today_sales)
    today_debt = sum(s.get('debt', 0) for s in today_sales)
    return jsonify({
        'totalSales': len(sales_data),
        'totalRevenue': sum(s['total'] for s in sales_data),
        'totalItems': sum(s['quantity'] for s in sales_data),
        'todaySales': sum(s['total'] for s in today_sales),
        'payments': {
            'totalMpesa': total_mpesa,
            'totalCash': total_cash,
            'totalDebt': total_debt,
            'todayMpesa': today_mpesa,
            'todayCash': today_cash,
            'todayDebt': today_debt
        }
    })

@app.route('/api/reports/sales')
@token_required
def sales_report():
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available'}), 500
    wb = openpyxl.Workbook()

    # Sheet 1: Detailed Sales (items)
    ws1 = wb.active
    ws1.title = "Sales Items"
    ws1.append(['Date', 'Transaction ID', 'Product', 'Quantity', 'Unit Price', 'Total'])
    for sale in sales_data:
        ws1.append([sale['date'], sale.get('transactionId', '-'), sale['product'], sale['quantity'], sale['unitPrice'], sale['total']])

    # Sheet 2: Payment Summary (per transaction)
    ws2 = wb.create_sheet("Payment Summary")
    ws2.append(['Date', 'Transaction ID', 'Items', 'Sale Total', 'M-Pesa', 'Cash', 'Debt', 'Payment Total'])

    # Group by transactionId
    transactions = {}
    for sale in sales_data:
        tid = sale.get('transactionId', sale['date'] + '-' + sale['product'])
        if tid not in transactions:
            transactions[tid] = {
                'date': sale['date'],
                'items': [],
                'total': 0,
                'mpesa': sale.get('mpesa', 0),
                'cash': sale.get('cash', 0),
                'debt': sale.get('debt', 0)
            }
        transactions[tid]['items'].append(f"{sale['product']} x{sale['quantity']}")
        transactions[tid]['total'] += sale['total']

    for tid, t in transactions.items():
        payment_total = t['mpesa'] + t['cash'] + t['debt']
        ws2.append([t['date'], tid, ', '.join(t['items']), t['total'], t['mpesa'], t['cash'], t['debt'], payment_total])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/api/reports/inventory')
@token_required
def inventory_report():
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available'}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    ws.append(['Product', 'Stock', 'Price', 'Value', 'Status'])
    for name, data in inventory_data.items():
        ws.append([name, data['stock'], data['price'], data['stock'] * data['price'], 'Low Stock' if data['stock'] < 5 else 'OK'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'inventory_report_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/api/reports/full')
@token_required
def full_report():
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available'}), 500
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(['Date', 'Transaction ID', 'Product', 'Quantity', 'Unit Price', 'Total', 'M-Pesa', 'Cash', 'Debt'])
    for sale in sales_data:
        ws1.append([sale['date'], sale.get('transactionId', '-'), sale['product'], sale['quantity'], sale['unitPrice'], sale['total'], sale.get('mpesa', 0), sale.get('cash', 0), sale.get('debt', 0)])
    ws2 = wb.create_sheet("Restocks")
    ws2.append(['Date', 'Product', 'Quantity', 'Unit Price', 'Total'])
    for restock in restocks_data:
        ws2.append([restock['date'], restock['product'], restock['quantity'], restock['unitPrice'], restock['total']])
    ws3 = wb.create_sheet("Inventory")
    ws3.append(['Product', 'Stock', 'Price', 'Value'])
    for name, data in inventory_data.items():
        ws3.append([name, data['stock'], data['price'], data['stock'] * data['price']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'full_report_{datetime.now().strftime("%Y%m%d")}.xlsx')

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Gerrit POS System</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }
        .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .login-container {
            background: white;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.2);
            max-width: 400px;
            margin: 100px auto;
        }
        .main-container {
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        .header {
            background: #333;
            color: white;
            padding: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 10px;
        }
        .nav {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }
        .nav button {
            background: #555;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 5px;
            cursor: pointer;
            transition: background 0.3s;
        }
        .nav button:hover, .nav button.active {
            background: #667eea;
        }
        .content { padding: 20px; }
        .hidden { display: none !important; }
        input, select {
            width: 100%;
            padding: 12px;
            margin: 8px 0;
            border: 1px solid #ddd;
            border-radius: 5px;
            font-size: 16px;
        }
        button {
            background: #667eea;
            color: white;
            border: none;
            padding: 12px 24px;
            border-radius: 5px;
            cursor: pointer;
            font-size: 16px;
            transition: background 0.3s;
        }
        button:hover { background: #5568d3; }
        .btn-danger { background: #e74c3c; }
        .btn-danger:hover { background: #c0392b; }
        .btn-success { background: #27ae60; }
        .btn-success:hover { background: #229954; }
        .btn-info { background: #17a2b8; }
        .btn-info:hover { background: #138496; }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th { background: #f8f9fa; font-weight: 600; }
        tr:hover { background: #f8f9fa; }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }
        .stat-value {
            font-size: 28px;
            font-weight: bold;
            margin: 10px 0;
        }
        .cart-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px;
            background: #f8f9fa;
            margin: 5px 0;
            border-radius: 5px;
        }
        .low-stock { color: #e74c3c; font-weight: bold; }
        .message {
            padding: 15px;
            margin: 10px 0;
            border-radius: 5px;
            display: none;
        }
        .message.success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        .message.error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        .search-box { margin-bottom: 20px; }
        .product-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 20px;
        }
        .product-card {
            border: 1px solid #ddd;
            padding: 15px;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.3s;
        }
        .product-card:hover {
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            transform: translateY(-2px);
        }
        .quantity-control {
            display: flex;
            align-items: center;
            gap: 10px;
            margin-top: 10px;
        }
        .quantity-control button {
            width: 30px;
            height: 30px;
            padding: 0;
            border-radius: 50%;
        }
        .cart-summary {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            margin-top: 20px;
        }
        .total {
            font-size: 24px;
            font-weight: bold;
            color: #667eea;
            margin-top: 10px;
        }
        .payment-section {
            margin-top: 20px;
            border-top: 2px solid #ddd;
            padding-top: 15px;
        }
        .payment-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10px;
            margin-top: 10px;
        }
        .payment-input label {
            font-size: 12px;
            color: #666;
            display: block;
            margin-bottom: 4px;
        }
        .payment-input input {
            margin: 0;
        }
        .payment-status {
            margin-top: 10px;
            font-size: 14px;
            color: #666;
        }
        .reports-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
            margin-top: 20px;
        }
        .report-card {
            border: 1px solid #ddd;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .report-card h3 {
            margin-bottom: 15px;
            color: #333;
        }
        .date-filter {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            align-items: center;
            flex-wrap: wrap;
        }
        .date-filter input {
            margin: 0;
            width: auto;
        }
    </style>
</head>
<body>
    <div class="container">
        <div id="loginSection" class="login-container">
            <h2 style="text-align: center; margin-bottom: 30px; color: #333;">Gerrit POS Login</h2>
            <div id="loginMessage" class="message"></div>
            <input type="text" id="username" placeholder="Username" value="admin">
            <input type="password" id="password" placeholder="Password">
            <button onclick="login()" style="width: 100%; margin-top: 10px;">Login</button>
        </div>

        <div id="appSection" class="main-container hidden">
            <div class="header">
                <h1>Gerrit POS System</h1>
                <div class="nav">
                    <button onclick="showTab('pos')" class="active" id="tab-pos">POS</button>
                    <button onclick="showTab('inventory')" id="tab-inventory">Inventory</button>
                    <button onclick="showTab('transactions')" id="tab-transactions">Transactions</button>
                    <button onclick="showTab('stats')" id="tab-stats">Stats</button>
                    <button onclick="showTab('reports')" id="tab-reports">Reports</button>
                    <button onclick="logout()" class="btn-danger">Logout</button>
                </div>
            </div>

            <div class="content">
                <div id="message" class="message"></div>

                <div id="posTab" class="tab-content">
                    <div class="search-box">
                        <input type="text" id="productSearch" placeholder="Search products..." onkeyup="filterProducts()">
                    </div>
                    <div class="pos-layout" style="display: flex; gap: 20px;">
                        <div class="products-area" style="flex: 1;">
                            <div class="product-grid" id="productGrid"></div>
                        </div>
                        <div class="cart-area" style="width: 320px; min-width: 320px;">
                            <div class="cart-summary" style="position: sticky; top: 20px;">
                                <h3>Cart</h3>
                                <div id="cartItems"></div>
                                <div class="total">Total: KES <span id="cartTotal">0</span></div>

                                <div style="margin: 15px 0;">
                                    <label style="font-size: 12px; color: #666;">Sale Date</label>
                                    <input type="date" id="saleDate" style="margin-top: 4px;">
                                </div>

                                <div class="payment-section">
                                    <h4>Payment Method</h4>
                                    <div class="payment-grid" style="grid-template-columns: 1fr;">
                                        <div class="payment-input">
                                            <label>M-Pesa (KES)</label>
                                            <input type="number" id="payMpesa" placeholder="0" min="0" step="0.01" oninput="validatePayment()">
                                        </div>
                                        <div class="payment-input">
                                            <label>Cash (KES)</label>
                                            <input type="number" id="payCash" placeholder="0" min="0" step="0.01" oninput="validatePayment()">
                                        </div>
                                        <div class="payment-input">
                                            <label>Debt (KES)</label>
                                            <input type="number" id="payDebt" placeholder="0" min="0" step="0.01" oninput="validatePayment()">
                                        </div>
                                    </div>
                                    <div class="payment-status">
                                        Payment Total: KES <span id="paymentTotal">0.00</span>
                                        <span id="paymentMatch" style="margin-left: 10px;"></span>
                                    </div>
                                </div>
                                <button onclick="checkout()" class="btn-success" style="width: 100%; margin-top: 15px;">Complete Sale</button>
                            </div>
                        </div>
                    </div>
                </div>

                <div id="inventoryTab" class="tab-content hidden">
                    <h2>Inventory Management</h2>
                    <div style="margin: 20px 0;">
                        <h3>Restock Product</h3>
                        <select id="restockProduct" onchange="updateRestockUnit()"></select>
                        <input type="number" id="restockQty" placeholder="Quantity" min="1">
                        <div id="restockUnitDisplay" style="font-size: 14px; color: #666; margin: 5px 0;">Unit: boxes</div>
                        <div style="margin: 10px 0;">
                            <label style="font-size: 12px; color: #666;">Restock Date</label>
                            <input type="date" id="restockDate" style="margin-top: 4px;">
                        </div>
                        <div id="restockPieceToggle" style="display: none; margin: 10px 0;">
                            <label style="font-size: 14px;"><input type="checkbox" id="restockByPiece"> Restock by pieces (instead of boxes)</label>
                        </div>
                        <button onclick="restock()">Restock</button>
                    </div>
                    <table id="inventoryTable">
                        <thead>
                            <tr><th>Product</th><th>Stock</th><th>Price (KES)</th><th>Status</th></tr>
                        </thead>
                        <tbody></tbody>
                    </table>
                </div>

                <div id="transactionsTab" class="tab-content hidden">
                    <h2>Transaction History</h2>
                    <div class="date-filter">
                        <label>From:</label>
                        <input type="date" id="dateFrom">
                        <label>To:</label>
                        <input type="date" id="dateTo">
                        <button onclick="loadTransactions()">Filter</button>
                        <button onclick="clearDateFilter()" class="btn-danger">Clear</button>
                    </div>
                    <table id="transactionsTable">
                        <thead>
                            <tr>
                                <th>Date</th><th>Type</th><th>Product</th><th>Qty</th>
                                <th>Total (KES)</th><th>M-Pesa</th><th>Cash</th><th>Debt</th>
                            </tr>
                        </thead>
                        <tbody></tbody>
                    </table>
                </div>

                <div id="statsTab" class="tab-content hidden">
                    <h2>Sales Statistics</h2>
                    <div class="stats-grid" id="statsGrid"></div>
                </div>

                <div id="reportsTab" class="tab-content hidden">
                    <h2>Download Reports</h2>
                    <div class="reports-grid">
                        <div class="report-card">
                            <h3>Sales Report</h3>
                            <p>All sales transactions with payment breakdown</p>
                            <button onclick="downloadReport('sales')" class="btn-info">Download Excel</button>
                        </div>
                        <div class="report-card">
                            <h3>Inventory Report</h3>
                            <p>Current stock levels and values</p>
                            <button onclick="downloadReport('inventory')" class="btn-info">Download Excel</button>
                        </div>
                        <div class="report-card">
                            <h3>Full Report</h3>
                            <p>Complete data: sales, restocks, inventory</p>
                            <button onclick="downloadReport('full')" class="btn-info">Download Excel</button>
                        </div>
                    </div>

                    <h2 style="margin-top: 40px;">Google Sheets Sync</h2>
                    <div class="reports-grid">
                        <div class="report-card">
                            <h3>Sync Status</h3>
                            <p id="syncStatusText">Checking...</p>
                            <button onclick="checkSyncStatus()" class="btn-info">Check Status</button>
                        </div>
                        <div class="report-card">
                            <h3>Manual Sync</h3>
                            <p>Push all data to Google Sheets now</p>
                            <button onclick="syncToSheets()" class="btn-success">Sync Now</button>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        let authToken = localStorage.getItem('pos_token');
        let inventory = {};
        let cart = {};
        let products = [];

        if (authToken) {
            checkAuth();
        }

        async function checkAuth() {
            try {
                const response = await fetch('/api/check-auth', {
                    headers: { 'Authorization': 'Bearer ' + authToken }
                });
                const data = await response.json();
                if (data.authenticated) {
                    showApp();
                } else {
                    localStorage.removeItem('pos_token');
                    showLogin();
                }
            } catch (error) {
                showLogin();
            }
        }

        function showLogin() {
            document.getElementById('loginSection').classList.remove('hidden');
            document.getElementById('appSection').classList.add('hidden');
        }

        function showApp() {
            document.getElementById('loginSection').classList.add('hidden');
            document.getElementById('appSection').classList.remove('hidden');
            loadProducts();
            loadInventory();
            setDefaultDates();
        }

        function setDefaultDates() {
            const today = new Date().toISOString().split('T')[0];
            const fromEl = document.getElementById('dateFrom');
            const toEl = document.getElementById('dateTo');
            const saleDateEl = document.getElementById('saleDate');
            const restockDateEl = document.getElementById('restockDate');
            if (fromEl) fromEl.value = today;
            if (toEl) toEl.value = today;
            if (saleDateEl) saleDateEl.value = today;
            if (restockDateEl) restockDateEl.value = today;
        }

        async function login() {
            const username = document.getElementById('username').value;
            const password = document.getElementById('password').value;

            try {
                const response = await fetch('/api/login', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ username, password })
                });

                const data = await response.json();
                if (data.success) {
                    authToken = data.token;
                    localStorage.setItem('pos_token', authToken);
                    showMessage('loginMessage', 'Login successful!', 'success');
                    showApp();
                } else {
                    showMessage('loginMessage', data.message, 'error');
                }
            } catch (error) {
                showMessage('loginMessage', 'Login failed', 'error');
            }
        }

        function logout() {
            fetch('/api/logout', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + authToken }
            });
            localStorage.removeItem('pos_token');
            authToken = null;
            showLogin();
        }

        async function loadProducts() {
            try {
                const response = await fetch('/api/products');
                products = await response.json();
                renderProductGrid();
                populateRestockSelect();
            } catch (error) {
                console.error('Failed to load products');
            }
        }

        async function loadInventory() {
            try {
                const response = await fetch('/api/inventory', {
                    headers: { 'Authorization': 'Bearer ' + authToken }
                });
                inventory = await response.json();
                renderInventoryTable();
            } catch (error) {
                console.error('Failed to load inventory');
            }
        }

        function renderProductGrid() {
            const grid = document.getElementById('productGrid');
            grid.innerHTML = '';

            products.forEach(product => {
                const isCigarette = product.pieces_per_box !== undefined;
                const pieceInfo = isCigarette ? `<p style="font-size: 12px; color: #666;">KES ${product.piece_price} per piece (${product.pieces_per_box} pieces/box)</p>` : '';
                const stockInfo = isCigarette ? 
                    `<p class="${product.stock < 1 ? 'low-stock' : ''}">${product.stock} boxes (${Math.floor(product.stock * product.pieces_per_box)} pieces)</p>` :
                    `<p class="${product.stock < 5 ? 'low-stock' : ''}">Stock: ${product.stock}</p>`;

                const card = document.createElement('div');
                card.className = 'product-card';
                card.innerHTML = `
                    <h4>${product.name}</h4>
                    <p>KES ${product.price} ${isCigarette ? 'per box' : ''}</p>
                    ${pieceInfo}
                    ${stockInfo}
                    <div class="quantity-control" onclick="event.stopPropagation()">
                        <button onclick="updateCart('${product.name}', -1)">-</button>
                        <span id="qty-${product.name}">0</span>
                        <button onclick="updateCart('${product.name}', 1)">+</button>
                    </div>
                    ${isCigarette ? `<div style="margin-top: 8px;"><label style="font-size: 12px;"><input type="checkbox" id="piece-${product.name}" onchange="togglePieceMode('${product.name}')"> Sell by piece</label></div>` : ''}
                `;
                grid.appendChild(card);
            });
        }

        let pieceMode = {}; // Track which products are in piece mode

        function togglePieceMode(product) {
            pieceMode[product] = document.getElementById(`piece-${product}`).checked;
            updateCartDisplay();
        }

        function filterProducts() {
            const search = document.getElementById('productSearch').value.toLowerCase();
            const cards = document.querySelectorAll('.product-card');
            cards.forEach(card => {
                const name = card.querySelector('h4').textContent.toLowerCase();
                card.style.display = name.includes(search) ? 'block' : 'none';
            });
        }

        function updateCart(product, change) {
            if (!cart[product]) cart[product] = 0;
            cart[product] += change;
            if (cart[product] < 0) cart[product] = 0;

            const productData = products.find(p => p.name === product);
            const isPiece = pieceMode[product];

            if (isPiece && productData.pieces_per_box) {
                // Piece mode: check against total pieces
                const totalPieces = Math.floor(productData.stock * productData.pieces_per_box);
                if (cart[product] > totalPieces) {
                    cart[product] = totalPieces;
                    showMessage('message', `Not enough pieces! Only ${totalPieces} available`, 'error');
                }
            } else {
                // Box mode: check against box stock
                if (cart[product] > productData.stock) {
                    cart[product] = productData.stock;
                    showMessage('message', 'Not enough stock!', 'error');
                }
            }

            document.getElementById(`qty-${product}`).textContent = cart[product];
            updateCartDisplay();
        }

        function updateCartDisplay() {
            const container = document.getElementById('cartItems');
            container.innerHTML = '';
            let total = 0;

            Object.entries(cart).forEach(([product, qty]) => {
                if (qty > 0) {
                    const productData = products.find(p => p.name === product);
                    const isPiece = pieceMode[product];

                    let itemTotal, unitLabel;
                    if (isPiece && productData.pieces_per_box) {
                        itemTotal = productData.piece_price * qty;
                        unitLabel = 'pieces';
                    } else {
                        itemTotal = productData.price * qty;
                        unitLabel = productData.pieces_per_box ? 'boxes' : '';
                    }
                    total += itemTotal;

                    const div = document.createElement('div');
                    div.className = 'cart-item';
                    div.innerHTML = `
                        <span>${product} x ${qty} ${unitLabel}</span>
                        <span>KES ${itemTotal.toFixed(2)}</span>
                    `;
                    container.appendChild(div);
                }
            });

            document.getElementById('cartTotal').textContent = total.toFixed(2);
            validatePayment();
        }

        function validatePayment() {
            const cartTotal = parseFloat(document.getElementById('cartTotal').textContent) || 0;
            const mpesa = parseFloat(document.getElementById('payMpesa').value) || 0;
            const cash = parseFloat(document.getElementById('payCash').value) || 0;
            const debt = parseFloat(document.getElementById('payDebt').value) || 0;
            const paymentTotal = mpesa + cash + debt;

            document.getElementById('paymentTotal').textContent = paymentTotal.toFixed(2);

            const matchEl = document.getElementById('paymentMatch');
            if (cartTotal > 0) {
                const diff = paymentTotal - cartTotal;
                if (Math.abs(diff) < 0.01) {
                    matchEl.innerHTML = '<span style="color: #27ae60; font-weight: bold;">✓ Balanced</span>';
                } else if (diff < 0) {
                    matchEl.innerHTML = `<span style="color: #e74c3c;">Short by KES ${Math.abs(diff).toFixed(2)}</span>`;
                } else {
                    matchEl.innerHTML = `<span style="color: #e74c3c;">Over by KES ${diff.toFixed(2)}</span>`;
                }
            } else {
                matchEl.textContent = '';
            }
        }

        async function checkout() {
            const items = Object.entries(cart)
                .filter(([_, qty]) => qty > 0)
                .map(([name, quantity]) => {
                    const product = products.find(p => p.name === name);
                    const isPiece = pieceMode[name];
                    const price = (isPiece && product.piece_price) ? product.piece_price : product.price;
                    return { name, quantity, price, unit_type: isPiece ? 'piece' : 'box' };
                });

            if (items.length === 0) {
                showMessage('message', 'Cart is empty!', 'error');
                return;
            }

            const total = items.reduce((sum, item) => sum + (item.price * item.quantity), 0);
            const mpesa = parseFloat(document.getElementById('payMpesa').value) || 0;
            const cash = parseFloat(document.getElementById('payCash').value) || 0;
            const debt = parseFloat(document.getElementById('payDebt').value) || 0;
            const totalInput = mpesa + cash + debt;
            const saleDate = document.getElementById('saleDate').value;

            if (Math.abs(totalInput - total) > 0.01) {
                showMessage('message', `Payment total (KES ${totalInput.toFixed(2)}) must equal sale total (KES ${total.toFixed(2)})`, 'error');
                return;
            }

            try {
                const response = await fetch('/api/sale', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                        'Authorization': 'Bearer ' + authToken
                    },
                    body: JSON.stringify({ 
                        items, 
                        payments: { mpesa, cash, debt },
                        date: saleDate || undefined
                    })
                });

                if (response.ok) {
                    showMessage('message', 'Sale completed!', 'success');
                    cart = {};
                    document.getElementById('payMpesa').value = '';
                    document.getElementById('payCash').value = '';
                    document.getElementById('payDebt').value = '';
                    loadProducts();
                    loadInventory();
                    renderProductGrid();
                } else {
                    const data = await response.json();
                    showMessage('message', data.error || 'Sale failed', 'error');
                }
            } catch (error) {
                showMessage('message', 'Sale failed', 'error');
            }
        }

        function renderInventoryTable() {
            const tbody = document.querySelector('#inventoryTable tbody');
            tbody.innerHTML = '';

            Object.entries(inventory).forEach(([name, data]) => {
                const isCigarette = data.pieces_per_box !== undefined;
                const stockDisplay = isCigarette ? 
                    `${data.stock} boxes (${Math.floor(data.stock * data.pieces_per_box)} pieces)` :
                    data.stock;
                const lowStock = isCigarette ? data.stock < 1 : data.stock < 5;
                const priceDisplay = isCigarette ? 
                    `KES ${data.price} (KES ${data.piece_price}/piece)` :
                    data.price;

                const row = document.createElement('tr');
                row.innerHTML = `
                    <td>${name}</td>
                    <td class="${lowStock ? 'low-stock' : ''}">${stockDisplay}</td>
                    <td>${priceDisplay}</td>
                    <td>${lowStock ? 'Low Stock' : 'OK'}</td>
                `;
                tbody.appendChild(row);
            });
        }

        function populateRestockSelect() {
            const select = document.getElementById('restockProduct');
            select.innerHTML = '';
            products.forEach(product => {
                const option = document.createElement('option');
                option.value = product.name;
                option.textContent = product.name;
                select.appendChild(option);
            });
            updateRestockUnit();
        }

        function updateRestockUnit() {
            const productName = document.getElementById('restockProduct').value;
            const product = products.find(p => p.name === productName);
            const unitDisplay = document.getElementById('restockUnitDisplay');
            const pieceToggle = document.getElementById('restockPieceToggle');

            if (product && product.pieces_per_box) {
                unitDisplay.textContent = `Unit: boxes (${product.pieces_per_box} pieces per box)`;
                pieceToggle.style.display = 'block';
            } else {
                unitDisplay.textContent = 'Unit: boxes/pieces';
                pieceToggle.style.display = 'none';
            }
        }

        async function restock() {
            const product = document.getElementById('restockProduct').value;
            const quantity = parseInt(document.getElementById('restockQty').value);
            const restockDate = document.getElementById('restockDate').value;
            const byPiece = document.getElementById('restockByPiece')?.checked || false;

            if (!quantity || quantity < 1) {
                showMessage('message', 'Invalid quantity', 'error');
                return;
            }

            try {
                const response = await fetch('/api/restock', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                        'Authorization': 'Bearer ' + authToken
                    },
                    body: JSON.stringify({ 
                        product, 
                        quantity, 
                        date: restockDate || undefined,
                        unit_type: byPiece ? 'piece' : 'box'
                    })
                });

                if (response.ok) {
                    showMessage('message', 'Restocked successfully!', 'success');
                    document.getElementById('restockQty').value = '';
                    loadInventory();
                    loadProducts();
                } else {
                    const data = await response.json();
                    showMessage('message', data.error || 'Restock failed', 'error');
                }
            } catch (error) {
                showMessage('message', 'Restock failed', 'error');
            }
        }

        function clearDateFilter() {
            setDefaultDates();
            loadTransactions();
        }

        async function loadTransactions() {
            try {
                const response = await fetch('/api/transactions', {
                    headers: { 'Authorization': 'Bearer ' + authToken }
                });
                let transactions = await response.json();

                const dateFrom = document.getElementById('dateFrom').value;
                const dateTo = document.getElementById('dateTo').value;

                if (dateFrom) {
                    transactions = transactions.filter(t => t.date >= dateFrom);
                }
                if (dateTo) {
                    transactions = transactions.filter(t => t.date <= dateTo);
                }

                const tbody = document.querySelector('#transactionsTable tbody');
                tbody.innerHTML = '';

                // Group transactions by transactionId to show payments once per sale
                const grouped = {};
                transactions.forEach(t => {
                    const key = t.transactionId || t.date + '-' + t.product;
                    if (!grouped[key]) {
                        grouped[key] = {
                            date: t.date,
                            type: t.type,
                            items: [],
                            total: 0,
                            mpesa: t.mpesa || 0,
                            cash: t.cash || 0,
                            debt: t.debt || 0
                        };
                    }
                    grouped[key].items.push(`${t.product} x${t.quantity}`);
                    grouped[key].total += t.total;
                });

                Object.values(grouped).forEach(g => {
                    const row = document.createElement('tr');
                    row.innerHTML = `
                        <td>${g.date}</td>
                        <td>${g.type}</td>
                        <td>${g.items.join(', ')}</td>
                        <td>-</td>
                        <td>KES ${g.total.toFixed(2)}</td>
                        <td>${g.mpesa ? 'KES ' + g.mpesa.toFixed(2) : '-'}</td>
                        <td>${g.cash ? 'KES ' + g.cash.toFixed(2) : '-'}</td>
                        <td>${g.debt ? 'KES ' + g.debt.toFixed(2) : '-'}</td>
                    `;
                    tbody.appendChild(row);
                });
            } catch (error) {
                console.error('Failed to load transactions');
            }
        }

        async function loadStats() {
            try {
                const response = await fetch('/api/stats', {
                    headers: { 'Authorization': 'Bearer ' + authToken }
                });
                const stats = await response.json();

                const grid = document.getElementById('statsGrid');
                const p = stats.payments || {};
                grid.innerHTML = `
                    <div class="stat-card">
                        <div>Total Sales</div>
                        <div class="stat-value">${stats.totalSales}</div>
                    </div>
                    <div class="stat-card">
                        <div>Total Revenue</div>
                        <div class="stat-value">KES ${stats.totalRevenue.toFixed(2)}</div>
                    </div>
                    <div class="stat-card">
                        <div>Items Sold</div>
                        <div class="stat-value">${stats.totalItems}</div>
                    </div>
                    <div class="stat-card">
                        <div>Today's Sales</div>
                        <div class="stat-value">KES ${stats.todaySales.toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #1e88e5 0%, #0d47a1 100%);">
                        <div>Total M-Pesa</div>
                        <div class="stat-value">KES ${(p.totalMpesa || 0).toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #43a047 0%, #1b5e20 100%);">
                        <div>Total Cash</div>
                        <div class="stat-value">KES ${(p.totalCash || 0).toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #e53935 0%, #b71c1c 100%);">
                        <div>Total Debt</div>
                        <div class="stat-value">KES ${(p.totalDebt || 0).toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #fb8c00 0%, #e65100 100%);">
                        <div>Today's M-Pesa</div>
                        <div class="stat-value">KES ${(p.todayMpesa || 0).toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #8e24aa 0%, #4a148c 100%);">
                        <div>Today's Cash</div>
                        <div class="stat-value">KES ${(p.todayCash || 0).toFixed(2)}</div>
                    </div>
                    <div class="stat-card" style="background: linear-gradient(135deg, #f4511e 0%, #bf360c 100%);">
                        <div>Today's Debt</div>
                        <div class="stat-value">KES ${(p.todayDebt || 0).toFixed(2)}</div>
                    </div>
                `;
            } catch (error) {
                console.error('Failed to load stats');
            }
        }

        function downloadReport(type) {
            const url = `/api/reports/${type}`;
            fetch(url, {
                headers: { 'Authorization': 'Bearer ' + authToken }
            })
            .then(response => {
                if (response.ok) {
                    return response.blob();
                }
                throw new Error('Report generation failed');
            })
            .then(blob => {
                const link = document.createElement('a');
                link.href = window.URL.createObjectURL(blob);
                link.download = `${type}_report_${new Date().toISOString().split('T')[0]}.xlsx`;
                link.click();
                showMessage('message', 'Report downloaded!', 'success');
            })
            .catch(error => {
                showMessage('message', error.message, 'error');
            });
        }

        async function checkSyncStatus() {
            try {
                const response = await fetch('/api/sync/status', {
                    headers: { 'Authorization': 'Bearer ' + authToken }
                });
                const data = await response.json();
                const statusText = document.getElementById('syncStatusText');
                if (data.connected) {
                    statusText.innerHTML = '<span style="color: #27ae60;">✓ Connected to Google Sheets</span>';
                } else {
                    statusText.innerHTML = '<span style="color: #e74c3c;">✗ Not connected. Check GOOGLE_CREDENTIALS env variable.</span>';
                }
            } catch (error) {
                showMessage('message', 'Failed to check sync status', 'error');
            }
        }

        async function syncToSheets() {
            try {
                const response = await fetch('/api/sync', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                        'Authorization': 'Bearer ' + authToken
                    },
                    body: JSON.stringify({ type: 'all' })
                });
                const data = await response.json();
                if (data.success) {
                    showMessage('message', 'Synced to Google Sheets!', 'success');
                } else {
                    showMessage('message', data.message || 'Sync failed', 'error');
                }
            } catch (error) {
                showMessage('message', 'Sync failed', 'error');
            }
        }

        function showTab(tab) {
            document.querySelectorAll('.tab-content').forEach(el => el.classList.add('hidden'));
            document.querySelectorAll('.nav button').forEach(el => el.classList.remove('active'));

            document.getElementById(tab + 'Tab').classList.remove('hidden');
            document.getElementById('tab-' + tab).classList.add('active');

            if (tab === 'inventory') loadInventory();
            if (tab === 'transactions') loadTransactions();
            if (tab === 'stats') loadStats();
        }

        function showMessage(elementId, text, type) {
            const el = document.getElementById(elementId);
            el.textContent = text;
            el.className = 'message ' + type;
            el.style.display = 'block';
            setTimeout(() => el.style.display = 'none', 3000);
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    app.run(debug=True)
